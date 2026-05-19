#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent / "lib"))
from export_legacy_19col_tsv import HEADERS, escape_tsv_cell, normalize_row, parse_markdown_table


def load_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".tsv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter="\t")
            headers = list(reader.fieldnames or [])
            if headers != HEADERS:
                raise ValueError("input TSV headers do not match legacy 19-column contract")
            return [dict(row) for row in reader]
    return [normalize_row(row) for row in parse_markdown_table(path.read_text(encoding="utf-8-sig"))]


def write_tsv(rows: list[dict[str, str]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    escaped_rows = [{key: escape_tsv_cell(value) for key, value in row.items()} for row in rows]
    with output.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS, delimiter="\t", quoting=csv.QUOTE_ALL, lineterminator="\n")
        writer.writeheader()
        writer.writerows(escaped_rows)


def render_cell(value: str) -> str:
    return (value or "").replace("\\n", "\n").replace("/n", "\n")


def write_basic_xlsx(rows: list[dict[str, str]], output: Path) -> None:
    from lib.minimal_xlsx import write_xlsx

    write_xlsx(HEADERS, [[row.get(header, "") for header in HEADERS] for row in rows], output, sheet_name="Legacy 19 Column Test Cases")


def write_formatted_xlsx(rows: list[dict[str, str]], output: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "API Test Cases"

    matrix = [HEADERS] + [[render_cell(row.get(header, "")) for header in HEADERS] for row in rows]
    for r_idx, row in enumerate(matrix, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    phase_fills = {
        "Method & Header": PatternFill("solid", fgColor="EAF3F8"),
        "Schema Validation": PatternFill("solid", fgColor="FFF2CC"),
        "Value, Business Logic, Cross Logic": PatternFill("solid", fgColor="E2F0D9"),
        "Business Logic": PatternFill("solid", fgColor="E2F0D9"),
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        fill = phase_fills.get(row[2].value or "")
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    widths = {
        "A": 24, "B": 28, "C": 30, "D": 36, "E": 48,
        "F": 58, "G": 64, "H": 76, "I": 76, "J": 14,
        "K": 12, "L": 12, "M": 12, "N": 18, "O": 18,
        "P": 18, "Q": 30, "R": 16, "S": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 36
    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 120

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="APITestCases", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    ws.add_table(table)

    for col_name in ["Manual Test Results Round 1", "Manual Test Results Round 2", "Automation Test Results"]:
        col_letter = get_column_letter(HEADERS.index(col_name) + 1)
        dv = DataValidation(type="list", formula1='"PASS,FAIL,BLOCKED,N/A"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    summary = wb.create_sheet("Summary")
    counts = Counter(row.get("Group Tests", "") for row in rows)
    summary_rows = [["Metric", "Value"], ["Total test cases", str(len(rows))]]
    summary_rows.extend([[group or "(blank)", str(count)] for group, count in sorted(counts.items())])
    summary_rows.append(["Format note", "Escaped TSV newlines are rendered as real Excel line breaks in manual-step cells."])
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in summary.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 72
    summary.sheet_view.showGridLines = False

    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export legacy 19-column testcase XLSX from Markdown or TSV")
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--tsv-output", type=Path)
    parser.add_argument("--formatted", action="store_true", help="Create a styled manual-execution workbook with filters, wrapping, summary, and real line breaks (default behavior)")
    parser.add_argument("--basic", action="store_true", help="Create a minimal XLSX without styling; formatted output is the default")
    args = parser.parse_args()

    rows = load_rows(args.input)
    if not rows:
        raise ValueError("input contains no testcase rows")
    if args.tsv_output:
        write_tsv(rows, args.tsv_output)
    if args.basic:
        write_basic_xlsx(rows, args.output)
    else:
        write_formatted_xlsx(rows, args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
