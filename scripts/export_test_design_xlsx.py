#!/usr/bin/env python3
"""Export Test Design Markdown to formatted XLSX."""
from __future__ import annotations

import argparse
import re
from pathlib import Path

def parse_test_design(md_content: str) -> list[dict[str, str]]:
    rows = []
    current_group = "General"
    
    # Simple markdown lines parsing
    lines = md_content.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Check for heading 2 (Group)
        if line.startswith("## ") and not line.startswith("## Control parameter") and not line.startswith("## Coverage Obligations"):
            current_group = line[3:].strip()
            i += 1
            continue
            
        # Check for TD Node (Heading 3)
        if line.startswith("### "):
            node_title = line[4:].strip()
            # Parse ID, Technique, Summary
            # Format: TD_P1_001 - [ST] - POST /v1/customer/validate...
            match = re.match(r"^(TD_[A-Z0-9_]+)\s*-\s*\[(.*?)\]\s*-\s*(.*)$", node_title, re.I)
            if match:
                td_id = match.group(1).strip()
                technique = match.group(2).strip()
                summary = match.group(3).strip()
            else:
                td_id = node_title.split("-")[0].strip()
                technique = ""
                summary = node_title
                
            steps = ""
            expected = ""
            source = ""
            
            # Parse details below heading
            i += 1
            while i < len(lines) and not lines[i].strip().startswith("###") and not lines[i].strip().startswith("## "):
                sub_line = lines[i].strip()
                if sub_line.startswith("- **Steps**:") or sub_line.startswith("- Steps:"):
                    steps = re.sub(r"^-\s*\*\*Steps\*\*:\s*|^-\s*Steps:\s*", "", sub_line).strip()
                elif sub_line.startswith("- **Expected**:") or sub_line.startswith("- Expected:"):
                    expected = re.sub(r"^-\s*\*\*Expected\*\*:\s*|^-\s*Expected:\s*", "", sub_line).strip()
                elif sub_line.startswith("- **Source**:") or sub_line.startswith("- Source:"):
                    source = re.sub(r"^-\s*\*\*Source\*\*:\s*|^-\s*Source:\s*", "", sub_line).strip()
                elif sub_line.startswith("-"):
                    # Check if steps/expected/source is empty or if we append to previous
                    cleaned = re.sub(r"^-\s*", "", sub_line).strip()
                    if cleaned.lower().startswith("steps:"):
                        steps = cleaned[6:].strip()
                    elif cleaned.lower().startswith("expected:"):
                        expected = cleaned[9:].strip()
                    elif cleaned.lower().startswith("source:"):
                        source = cleaned[7:].strip()
                i += 1
                
            rows.append({
                "Group": current_group,
                "Test Design ID": td_id,
                "Technique": technique,
                "Test Condition Summary": summary,
                "Steps": steps,
                "Expected": expected,
                "Source": source
            })
            continue
        i += 1
    return rows

def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Design"
    ws.sheet_view.showGridLines = False
    
    headers = ["Function/Group", "Test Design ID", "Technique", "Test Condition Summary", "Steps", "Expected", "Source"]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        
    # Write data
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row_idx, 1, row["Group"])
        ws.cell(row_idx, 2, row["Test Design ID"])
        ws.cell(row_idx, 3, row["Technique"])
        ws.cell(row_idx, 4, row["Test Condition Summary"])
        ws.cell(row_idx, 5, row["Steps"])
        ws.cell(row_idx, 6, row["Expected"])
        ws.cell(row_idx, 7, row["Source"])
        
    # Styling
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        
    phase_fills = {
        "Method & Header": PatternFill("solid", fgColor="EAF3F8"),
        "Schema Validation": PatternFill("solid", fgColor="FFF2CC"),
        "Value, Business Logic, Cross Logic": PatternFill("solid", fgColor="E2F0D9"),
        "Business Logic": PatternFill("solid", fgColor="E2F0D9"),
    }
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        group_val = row[0].value or ""
        fill = None
        for key, p_fill in phase_fills.items():
            if key in group_val:
                fill = p_fill
                break
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
                
    widths = {
        "A": 28, "B": 20, "C": 14, "D": 48, "E": 64, "F": 64, "G": 40
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 28
    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 80
        
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="TestDesignTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    ws.add_table(table)
    
    wb.save(output_path)

def main() -> int:
    parser = argparse.ArgumentParser(description="Export Test Design Markdown to XLSX")
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    
    if not args.input.exists():
        print(f"Error: input file {args.input} does not exist", file=sys.stderr)
        return 1
        
    content = args.input.read_text(encoding="utf-8-sig")
    rows = parse_test_design(content)
    if not rows:
        print("Warning: parsed 0 rows from test design", file=sys.stderr)
    write_xlsx(rows, args.output)
    print(f"Successfully exported test design to {args.output}")
    return 0

if __name__ == "__main__":
    import sys
    raise SystemExit(main())
