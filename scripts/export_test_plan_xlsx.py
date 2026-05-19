#!/usr/bin/env python3
"""Export Test Plan Markdown to formatted XLSX."""
from __future__ import annotations

import argparse
import re
from pathlib import Path

def parse_test_plan(md_content: str) -> dict[str, object]:
    # Extract metadata at top
    metadata = {}
    metadata_re = re.compile(r"^\*\*(?P<key>[^*]+)\*\*\s*:\s*(?P<value>.*)$")
    for line in md_content.splitlines():
        match = metadata_re.match(line.strip())
        if match:
            metadata[match.group("key").strip()] = match.group("value").strip()
            
    # Extract sections
    sections = {}
    # Find all ## Headings
    headings = list(re.finditer(r"^##\s+(.+?)\s*$", md_content, re.MULTILINE))
    for idx, match in enumerate(headings):
        title = match.group(1).strip()
        start = match.end()
        end = headings[idx+1].start() if idx + 1 < len(headings) else len(md_content)
        content = md_content[start:end].strip()
        
        # Parse content: check if it contains a table
        if "|" in content:
            tables = []
            table = []
            for line in [*content.splitlines(), ""]:
                stripped = line.strip()
                if stripped.startswith("|"):
                    cells = [cell.strip() for cell in stripped.strip("|").split("|")]
                    if all(set(cell) <= {"-", ":", " "} for cell in cells):
                        continue
                    table.append(cells)
                    continue
                if len(table) >= 2:
                    tables.append(table)
                table = []
            if tables:
                sections[title] = {"type": "table", "data": tables[0]}
                continue
                
        # Parse content as text list/lines
        lines = [line.strip() for line in content.splitlines() if line.strip()]
        sections[title] = {"type": "text", "data": lines}
        
    return {"metadata": metadata, "sections": sections}

def write_xlsx(plan_data: dict[str, object], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Common styles
    blue_header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    regular_font = Font()
    thin_border = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Sheet structure definition supporting both EN and VI headers
    sheet_mapping = {
        "Tổng quan & Lịch trình": [
            "Scope In", "Phạm vi áp dụng",
            "Scope Out", "Phạm vi loại trừ",
            "Entry Criteria", "Điều kiện bắt đầu",
            "Exit Criteria", "Điều kiện kết thúc",
            "Schedule / Milestones", "Lịch trình / Mốc thời gian"
        ],
        "Căn cứ & Độ bao phủ": [
            "Source Baseline", "Tài liệu căn cứ",
            "Requirement Baseline", "Yêu cầu căn cứ",
            "Coverage Strategy", "Chiến lược độ bao phủ"
        ],
        "Tài nguyên & Môi trường": [
            "Test Levels / Phases", "Mức độ / Giai đoạn kiểm thử",
            "Deliverables", "Sản phẩm bàn giao",
            "Roles and Responsibilities", "Vai trò và Trách nhiệm",
            "Environment / Test Data / Dependencies", "Môi trường / Dữ liệu / Phụ thuộc",
            "Risks and Mitigations", "Rủi ro và Giải pháp",
            "Open Questions", "Câu hỏi cần làm rõ"
        ]
    }
    
    # Helper to write a section (table or text list)
    def write_section(ws, current_row: int, title: str, section: dict) -> int:
        # Write Section Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        header_cell = ws.cell(current_row, 1, title)
        header_cell.font = Font(size=12, bold=True, color="1F4E78")
        header_cell.fill = section_fill
        ws.row_dimensions[current_row].height = 26
        current_row += 1
        
        if section["type"] == "table":
            table_data = section["data"]
            headers = table_data[0]
            # Write Table Headers
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(current_row, col_idx, h)
                c.fill = blue_header_fill
                c.font = white_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = border
            ws.row_dimensions[current_row].height = 26
            current_row += 1
            
            # Write Table Rows
            for row in table_data[1:]:
                for col_idx, val in enumerate(row, start=1):
                    c = ws.cell(current_row, col_idx, val)
                    c.border = border
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[current_row].height = 42
                current_row += 1
        else:
            # Write Text/List lines
            for line in section["data"]:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                c = ws.cell(current_row, 1, line)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = Border(bottom=thin_border)
                ws.row_dimensions[current_row].height = 22
                current_row += 1
                
        return current_row + 2

    # 1. Tổng quan & Lịch trình Sheet
    ws1 = wb.create_sheet(title="Tổng quan & Lịch trình")
    ws1.sheet_view.showGridLines = True
    ws1.cell(1, 1, "KẾ HOẠCH KIỂM THỬ - TỔNG QUAN & LỊCH TRÌNH").font = Font(size=16, bold=True, color="1F4E78")
    
    # Metadata Block
    ws1.cell(3, 1, "Thông tin chung").font = Font(size=12, bold=True)
    m_row = 4
    for key, value in plan_data["metadata"].items():
        cell_k = ws1.cell(m_row, 1, key)
        cell_k.font = bold_font
        cell_k.border = border
        cell_k.fill = PatternFill("solid", fgColor="F2F2F2")
        
        cell_v = ws1.cell(m_row, 2, value)
        cell_v.font = regular_font
        cell_v.border = border
        m_row += 1
        
    # Write mapped sections
    row1 = m_row + 2
    for title in sheet_mapping["Tổng quan & Lịch trình"]:
        if title in plan_data["sections"]:
            row1 = write_section(ws1, row1, title, plan_data["sections"][title])
            
    # Set default column widths for sheet 1
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 40
    ws1.column_dimensions["E"].width = 25
    ws1.column_dimensions["F"].width = 25

    # 2. Căn cứ & Độ bao phủ Sheet
    ws2 = wb.create_sheet(title="Căn cứ & Độ bao phủ")
    ws2.sheet_view.showGridLines = True
    ws2.cell(1, 1, "KẾ HOẠCH KIỂM THỬ - CĂN CỨ & ĐỘ BAO PHỦ").font = Font(size=16, bold=True, color="1F4E78")
    
    # Trace ID ref
    tp_id = plan_data["metadata"].get("Test Plan ID", plan_data["metadata"].get("Mã Kế Hoạch Kiểm Thử", "N/A"))
    ws2.cell(3, 1, "Mã kế hoạch kiểm thử:").font = bold_font
    ws2.cell(3, 2, tp_id).font = regular_font
    
    row2 = 5
    for title in sheet_mapping["Căn cứ & Độ bao phủ"]:
        if title in plan_data["sections"]:
            row2 = write_section(ws2, row2, title, plan_data["sections"][title])
            
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 65
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 25
    ws2.column_dimensions["F"].width = 25

    # 3. Tài nguyên & Môi trường Sheet
    ws3 = wb.create_sheet(title="Tài nguyên & Môi trường")
    ws3.sheet_view.showGridLines = True
    ws3.cell(1, 1, "KẾ HOẠCH KIỂM THỬ - TÀI NGUYÊN & MÔI TRƯỜNG").font = Font(size=16, bold=True, color="1F4E78")
    
    ws3.cell(3, 1, "Mã kế hoạch kiểm thử:").font = bold_font
    ws3.cell(3, 2, tp_id).font = regular_font
    
    row3 = 5
    for title in sheet_mapping["Tài nguyên & Môi trường"]:
        if title in plan_data["sections"]:
            row3 = write_section(ws3, row3, title, plan_data["sections"][title])
            
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 35
    ws3.column_dimensions["D"].width = 45
    ws3.column_dimensions["E"].width = 35
    ws3.column_dimensions["F"].width = 35

    wb.save(output_path)

def main() -> int:
    parser = argparse.ArgumentParser(description="Export Test Plan Markdown to XLSX")
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    
    if not args.input.exists():
        print(f"Error: input file {args.input} does not exist", file=sys.stderr)
        return 1
        
    content = args.input.read_text(encoding="utf-8-sig")
    plan_data = parse_test_plan(content)
    write_xlsx(plan_data, args.output)
    print(f"Successfully exported test plan to {args.output}")
    return 0

if __name__ == "__main__":
    import sys
    raise SystemExit(main())
