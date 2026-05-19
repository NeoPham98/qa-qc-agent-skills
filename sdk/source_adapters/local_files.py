from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from source_manifest import SourceItem, source_item_from_path

SUPPORTED_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".txt", ".md", ".csv", ".tsv", ".xlsx", ".json", ".yaml", ".yml", ".feature"
}


def expand_sources(paths: Iterable[Path], root: Path | None = None) -> list[SourceItem]:
    files: list[Path] = []
    for path in paths:
        if path.is_dir():
            files.extend(sorted(p for p in path.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS))
        elif path.is_file():
            files.append(path)
        else:
            raise FileNotFoundError(f"Source does not exist: {path}")
    items = [source_item_from_path(f"source-{idx + 1}", path, root=root) for idx, path in enumerate(files)]
    for item in items:
        path = Path(item.local_path) if item.local_path else Path(item.original_locator)
        if not path.is_absolute() and root is not None:
            path = root / path
        item.fingerprint.detected_headers = detect_headers(path)
    return items


def detect_headers(path: Path) -> list[str]:
    ext = path.suffix.lower()
    if ext in {".csv", ".tsv"}:
        delimiter = "\t" if ext == ".tsv" else ","
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    if any(cell.strip() for cell in row):
                        return [cell.strip() for cell in row]
        except UnicodeDecodeError:
            return []
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 10), values_only=True):
                vals = [str(v).strip() if v is not None else "" for v in row]
                if any(vals):
                    return vals
        except Exception:
            return []
    return []


def sample_text(path: Path, limit: int = 8000) -> str:
    if path.suffix.lower() not in {".txt", ".md", ".csv", ".tsv", ".json", ".yaml", ".yml", ".feature"}:
        return ""
    try:
        return path.read_text(encoding="utf-8", errors="ignore")[:limit]
    except OSError:
        return ""
