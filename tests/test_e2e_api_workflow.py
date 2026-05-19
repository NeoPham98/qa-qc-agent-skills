from __future__ import annotations

import json
from openpyxl import load_workbook
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
SDK = ROOT / "sdk"
EXAMPLE = ROOT / "examples" / "e2e" / "api-cctg"
sys.path.insert(0, str(SDK))

from source_manifest import SourceFingerprint, SourceItem, SourceManifest


FINAL_OUTPUTS = [
    "API_OperationInventory.md",
    "API_TestDesign.md",
    "TestCaseSource.md",
    "Legacy19TestCase.generated.tsv",
    "Legacy19TestCase.generated.xlsx",
    "OutputReview.md",
    "SupervisorApproval.md",
]

FORBIDDEN_SECRET_MARKERS = [
    "Bearer ",
    "password=",
    "PASSWORD=",
    "TestLink key",
    "10.0.",
    "192.168.",
]


def run_script(script: str, *args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(SCRIPTS / script), *args],
        capture_output=True,
        text=True,
        check=False,
    )


def run_closed_loop(*args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(SCRIPTS / "run_closed_loop.py"), *args],
        capture_output=True,
        text=True,
        check=False,
    )


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def write_manifest(path: Path, output_dir: Path) -> None:
    manifest = SourceManifest(
        schema_version="1.0",
        run_id="api-e2e-test",
        user_prompt="Sinh testcase API BIDV 19 cột cho PAYGATES CCTG",
        output_directory=str(output_dir),
        workflow_pack="default",
        sources=[
            SourceItem(
                id="api-cctg-source",
                kind="local_file",
                original_locator="examples/e2e/api-cctg/API_TestDesign.md",
                local_path="examples/e2e/api-cctg/API_TestDesign.md",
                extension=".md",
                fingerprint=SourceFingerprint(candidate_roles=["api_spec"]),
            )
        ],
    )
    manifest.write(path)


def copy_seed_outputs(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for name in ["API_OperationInventory.md", "API_TestDesign.md", "TestCaseSource.md", "OutputReview.md", "SupervisorApproval.md"]:
        shutil.copy2(EXAMPLE / name, output_dir / name)


def assert_no_raw_secret(output_dir: Path) -> None:
    for name in FINAL_OUTPUTS:
        path = output_dir / name
        if path.suffix.lower() == ".xlsx":
            continue
        text = path.read_text(encoding="utf-8", errors="replace")
        for marker in FORBIDDEN_SECRET_MARKERS:
            assert_true(marker not in text, f"{path} contains forbidden marker {marker!r}")


def test_api_e2e_workflow_generates_valid_approved_artifacts() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        output_dir = tmp_path / "out"
        manifest = tmp_path / "source_manifest.json"
        copy_seed_outputs(output_dir)
        write_manifest(manifest, output_dir)

        export_result = run_script(
            "export_legacy_19col_xlsx.py",
            str(output_dir / "TestCaseSource.md"),
            str(output_dir / "Legacy19TestCase.generated.xlsx"),
            "--tsv-output",
            str(output_dir / "Legacy19TestCase.generated.tsv"),
            "--formatted",
        )
        assert_true(export_result.returncode == 0, export_result.stdout + export_result.stderr)
        workbook = load_workbook(output_dir / "Legacy19TestCase.generated.xlsx")
        worksheet = workbook["API Test Cases"]
        assert_true(worksheet["H2"].alignment.wrap_text is True, "formatted workbook must wrap test steps")
        multiline_values = [cell.value for row in worksheet.iter_rows(min_row=2) for cell in row if isinstance(cell.value, str) and "\n" in cell.value]
        assert_true(bool(multiline_values), "formatted workbook must render real line breaks when source cells contain logical newlines")
        assert_true(all("\\n" not in value and "/n" not in value for value in multiline_values), "formatted workbook must not contain literal newline markers")
        assert_true("Summary" in workbook.sheetnames, "formatted workbook must include Summary sheet")

        td_result = run_script("validate_test_design.py", str(output_dir / "API_TestDesign.md"), "--type", "api")
        assert_true(td_result.returncode == 0, td_result.stdout + td_result.stderr)

        td_specificity_result = run_script("validate_api_td_specificity.py", str(output_dir / "API_TestDesign.md"))
        assert_true(td_specificity_result.returncode == 0, td_specificity_result.stdout + td_specificity_result.stderr)

        tc_specificity_result = run_script("validate_api_tc_specificity.py", str(output_dir / "Legacy19TestCase.generated.tsv"))
        assert_true(tc_specificity_result.returncode == 0, tc_specificity_result.stdout + tc_specificity_result.stderr)

        tc_granularity_result = run_script(
            "validate_testcase_granularity.py",
            str(output_dir / "Legacy19TestCase.generated.tsv"),
            "--profile",
            "api_legacy_19_column_testcase",
        )
        assert_true(tc_granularity_result.returncode == 0, tc_granularity_result.stdout + tc_granularity_result.stderr)

        tc_coverage_result = run_script(
            "validate_testcase_coverage.py",
            str(output_dir / "Legacy19TestCase.generated.tsv"),
            "--profile",
            "api_legacy_19_column_testcase",
        )
        assert_true(tc_coverage_result.returncode == 0, tc_coverage_result.stdout + tc_coverage_result.stderr)

        contract_result = run_script(
            "validate_testcase_contract.py",
            str(output_dir / "Legacy19TestCase.generated.tsv"),
            "--profile",
            "legacy_19_column_testcase",
        )
        assert_true(contract_result.returncode == 0, contract_result.stdout + contract_result.stderr)

        legacy_result = run_script("validate_legacy_19col_tsv.py", str(output_dir / "Legacy19TestCase.generated.tsv"))
        assert_true(legacy_result.returncode == 0, legacy_result.stdout + legacy_result.stderr)

        for name in FINAL_OUTPUTS:
            assert_true((output_dir / name).exists(), f"missing final output {name}")
        assert_no_raw_secret(output_dir)

        approved_root = ROOT / "artifacts" / "default" / "approved" / "PAYGATES_E2E" / "Squad_Base" / "CCTG_API"
        reviewed_root = ROOT / "artifacts" / "default" / "reviewed" / "PAYGATES_E2E" / "Squad_Base" / "CCTG_API"
        archive_root = ROOT / "artifacts" / "default" / "archive" / "PAYGATES_E2E" / "Squad_Base" / "CCTG_API"
        rejected_root = ROOT / "artifacts" / "default" / "rejected" / "PAYGATES_E2E" / "Squad_Base" / "CCTG_API"
        for root in [approved_root, reviewed_root, archive_root, rejected_root]:
            shutil.rmtree(root, ignore_errors=True)

        try:
            loop_result = run_closed_loop(
                "--manifest",
                str(manifest),
                "--route-id",
                "api_spec_to_testcase",
                "--output-dir",
                str(output_dir),
                "--project",
                "PAYGATES_E2E",
                "--squad",
                "Squad_Base",
                "--epic",
                "CCTG_API",
            )
            assert_true(loop_result.returncode == 0, loop_result.stdout + loop_result.stderr)
            payload = json.loads(loop_result.stdout)
            assert_true(payload["status"] == "passed", str(payload))
            assert_true(payload["lifecycle_state"] == "approved", str(payload))
            assert_true(Path(payload["approved_artifact_path"]).exists(), "approved artifact missing")
            assert_true(Path(payload["reviewed_artifact_path"]).exists(), "reviewed artifact missing")
            assert_true((output_dir / "published-artifact-manifest.yml").exists(), "published manifest missing")
            assert_true((output_dir / "closed-loop-state.json").exists(), "closed-loop state missing")
        finally:
            for root in [approved_root, reviewed_root, archive_root, rejected_root]:
                shutil.rmtree(root, ignore_errors=True)


if __name__ == "__main__":
    test_api_e2e_workflow_generates_valid_approved_artifacts()
    print("test_e2e_api_workflow: PASS")
