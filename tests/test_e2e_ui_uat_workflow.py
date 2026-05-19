from __future__ import annotations

import csv
import json
from openpyxl import load_workbook
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
SDK = ROOT / "sdk"
EXAMPLE = ROOT / "examples" / "e2e" / "ui-uat"
sys.path.insert(0, str(SDK))

from source_manifest import SourceFingerprint, SourceItem, SourceManifest


LEGACY_OUTPUTS = [
    "UI_TestDesign.md",
    "TestCaseSource.md",
    "Legacy19TestCase.generated.tsv",
    "Legacy19TestCase.generated.xlsx",
    "OutputReview.md",
    "SupervisorApproval.md",
]
UAT_HEADERS = [
    "Test Case ID",
    "Test Case Name",
    "Precondition",
    "Test Steps",
    "Test Data",
    "Expected Result",
    "Priority",
    "Test Type",
    "Module",
    "Requirement ID",
    "Execution Result",
    "Actual Result",
    "Bug ID",
    "Tester",
    "Execution Date",
    "Notes",
]


def run_script(script: str, *args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(SCRIPTS / script), *args],
        capture_output=True,
        text=True,
        check=False,
    )


def run_closed_loop(*args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, str(SCRIPTS / "run_closed_loop.py"), *args],
        capture_output=True,
        text=True,
        check=False,
    )


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def copy_common_reviews(output_dir: Path) -> None:
    shutil.copy2(EXAMPLE / "OutputReview.md", output_dir / "OutputReview.md")
    shutil.copy2(EXAMPLE / "SupervisorApproval.md", output_dir / "SupervisorApproval.md")


def write_manifest(path: Path, output_dir: Path, role: str, prompt: str, source_name: str) -> None:
    manifest = SourceManifest(
        schema_version="1.0",
        run_id=f"{role}-e2e-test",
        user_prompt=prompt,
        output_directory=str(output_dir),
        workflow_pack="default",
        sources=[
            SourceItem(
                id=f"{role}-source",
                kind="local_file",
                original_locator=f"examples/e2e/ui-uat/{source_name}",
                local_path=f"examples/e2e/ui-uat/{source_name}",
                extension=".md",
                fingerprint=SourceFingerprint(candidate_roles=[role]),
            )
        ],
    )
    manifest.write(path)


def write_uat_tsv(source: Path, output: Path) -> None:
    rows = []
    for line in source.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped.startswith("|") or "---" in stripped:
            continue
        rows.append([cell.strip() for cell in stripped.strip("|").split("|")])
    header = rows[0]
    data_rows = [dict(zip(header, row)) for row in rows[1:] if len(row) == len(header)]
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=UAT_HEADERS, delimiter="\t", quoting=csv.QUOTE_ALL, lineterminator="\n")
        writer.writeheader()
        writer.writerows(data_rows)


def assert_no_raw_secret(output_dir: Path, names: list[str]) -> None:
    for name in names:
        path = output_dir / name
        if path.suffix.lower() == ".xlsx":
            continue
        text = path.read_text(encoding="utf-8", errors="replace")
        for marker in ["Bearer ", "password=", "PASSWORD=", "10.0.", "192.168."]:
            assert_true(marker not in text, f"{path} contains forbidden marker {marker!r}")


def run_publish_check(output_dir: Path, manifest: Path, route_id: str, epic: str) -> None:
    approved_root = ROOT / "artifacts" / "default" / "approved" / "PAYGATES_E2E" / "Squad_UI" / epic
    reviewed_root = ROOT / "artifacts" / "default" / "reviewed" / "PAYGATES_E2E" / "Squad_UI" / epic
    rejected_root = ROOT / "artifacts" / "default" / "rejected" / "PAYGATES_E2E" / "Squad_UI" / epic
    archive_root = ROOT / "artifacts" / "default" / "archive" / "PAYGATES_E2E" / "Squad_UI" / epic
    for root in [approved_root, reviewed_root, rejected_root, archive_root]:
        shutil.rmtree(root, ignore_errors=True)
    try:
        result = run_closed_loop(
            "--manifest",
            str(manifest),
            "--route-id",
            route_id,
            "--output-dir",
            str(output_dir),
            "--project",
            "PAYGATES_E2E",
            "--squad",
            "Squad_UI",
            "--epic",
            epic,
        )
        assert_true(result.returncode == 0, result.stdout + result.stderr)
        payload = json.loads(result.stdout)
        assert_true(payload["status"] == "passed", str(payload))
        assert_true(payload["lifecycle_state"] == "approved", str(payload))
        assert_true(Path(payload["approved_artifact_path"]).exists(), "approved artifact missing")
        assert_true(Path(payload["reviewed_artifact_path"]).exists(), "reviewed artifact missing")
    finally:
        for root in [approved_root, reviewed_root, rejected_root, archive_root]:
            shutil.rmtree(root, ignore_errors=True)


def test_ui_and_uat_e2e_workflows_generate_valid_artifacts() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        ui_out = tmp_path / "ui-out"
        ui_out.mkdir()
        shutil.copy2(EXAMPLE / "UI_TestDesign.md", ui_out / "UI_TestDesign.md")
        shutil.copy2(EXAMPLE / "UI_TestCaseSource.md", ui_out / "TestCaseSource.md")
        copy_common_reviews(ui_out)
        ui_manifest = tmp_path / "ui_manifest.json"
        write_manifest(ui_manifest, ui_out, "ui_spec", "Sinh testcase UI BIDV 19 cột", "UI_TestDesign.md")

        ui_td_result = run_script("validate_test_design.py", str(ui_out / "UI_TestDesign.md"), "--type", "ui")
        assert_true(ui_td_result.returncode == 0, ui_td_result.stdout + ui_td_result.stderr)
        ui_export_result = run_script(
            "export_legacy_19col_xlsx.py",
            str(ui_out / "TestCaseSource.md"),
            str(ui_out / "Legacy19TestCase.generated.xlsx"),
            "--tsv-output",
            str(ui_out / "Legacy19TestCase.generated.tsv"),
            "--formatted",
        )
        assert_true(ui_export_result.returncode == 0, ui_export_result.stdout + ui_export_result.stderr)
        workbook = load_workbook(ui_out / "Legacy19TestCase.generated.xlsx")
        worksheet = workbook["API Test Cases"]
        assert_true(worksheet["H2"].alignment.wrap_text is True, "formatted workbook must wrap test steps")
        assert_true("Summary" in workbook.sheetnames, "formatted workbook must include Summary sheet")
        ui_contract_result = run_script(
            "validate_testcase_contract.py",
            str(ui_out / "Legacy19TestCase.generated.tsv"),
            "--profile",
            "legacy_19_column_testcase",
        )
        assert_true(ui_contract_result.returncode == 0, ui_contract_result.stdout + ui_contract_result.stderr)
        ui_granularity_result = run_script(
            "validate_testcase_granularity.py",
            str(ui_out / "Legacy19TestCase.generated.tsv"),
            "--profile",
            "ui_legacy_19_column_testcase",
        )
        assert_true(ui_granularity_result.returncode == 0, ui_granularity_result.stdout + ui_granularity_result.stderr)
        assert_no_raw_secret(ui_out, LEGACY_OUTPUTS)
        run_publish_check(ui_out, ui_manifest, "ui_source_to_testcase", "UI_FLOW")

        uat_out = tmp_path / "uat-out"
        uat_out.mkdir()
        shutil.copy2(EXAMPLE / "UAT_TestCaseSource.md", uat_out / "UAT_TestCaseSource.md")
        copy_common_reviews(uat_out)
        write_uat_tsv(uat_out / "UAT_TestCaseSource.md", uat_out / "UAT_TestCase.generated.tsv")
        uat_manifest = tmp_path / "uat_manifest.json"
        write_manifest(uat_manifest, uat_out, "urd", "Sinh testcase UAT BIDV 16 cột", "UAT_TestCaseSource.md")

        uat_contract_result = run_script(
            "validate_testcase_contract.py",
            str(uat_out / "UAT_TestCase.generated.tsv"),
            "--profile",
            "uat_16_column_testcase",
        )
        assert_true(uat_contract_result.returncode == 0, uat_contract_result.stdout + uat_contract_result.stderr)
        uat_granularity_result = run_script(
            "validate_testcase_granularity.py",
            str(uat_out / "UAT_TestCase.generated.tsv"),
            "--profile",
            "uat_16_column_testcase",
        )
        assert_true(uat_granularity_result.returncode == 0, uat_granularity_result.stdout + uat_granularity_result.stderr)
        assert_no_raw_secret(uat_out, ["UAT_TestCaseSource.md", "UAT_TestCase.generated.tsv", "OutputReview.md", "SupervisorApproval.md"])
        run_publish_check(uat_out, uat_manifest, "urd_to_uat_testcase", "UAT_FLOW")


if __name__ == "__main__":
    test_ui_and_uat_e2e_workflows_generate_valid_artifacts()
    print("test_e2e_ui_uat_workflow: PASS")
